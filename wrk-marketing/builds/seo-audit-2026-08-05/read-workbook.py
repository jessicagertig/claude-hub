"""Read the SEO audit workbook, surfacing formatting that carries meaning.

Usage:
  wb.py survey            list tabs + what formatting is present in each
  wb.py <tab substring>   dump one tab: values + non-default formatting per cell
"""
import sys
import openpyxl

PATH = '/Users/jessica/claude-hub/wrk-marketing/builds/seo-audit-2026-08-05/Polymer-Technical-SEO-Audit_MakeReality.xlsx'


def fmt_marks(cell):
    """Return the formatting on a cell that could carry meaning."""
    marks = []
    f = cell.font
    if f.bold:
        marks.append('bold')
    if f.strike:
        marks.append('STRIKE')
    if f.italic:
        marks.append('italic')
    if f.color is not None and f.color.type == 'rgb' and f.color.rgb not in (None, 'FF000000', '00000000'):
        marks.append(f'color:{f.color.rgb}')
    fill = cell.fill
    if fill is not None and fill.fill_type == 'solid':
        rgb = getattr(fill.fgColor, 'rgb', None)
        if rgb not in (None, '00000000', 'FFFFFFFF'):
            marks.append(f'fill:{rgb}')
    if cell.comment is not None:
        marks.append(f'comment:{cell.comment.text.strip()[:80]}')
    if cell.hyperlink is not None:
        marks.append(f'link:{cell.hyperlink.target}')
    return marks


wb = openpyxl.load_workbook(PATH, data_only=True)
arg = sys.argv[1] if len(sys.argv) > 1 else 'survey'

if arg == 'survey':
    for name in wb.sheetnames:
        ws = wb[name]
        seen, cells = set(), 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cells += 1
                for m in fmt_marks(cell):
                    seen.add(m.split(':')[0])
        print(f'{name!r}  rows={ws.max_row} cols={ws.max_column} filled={cells}  formatting={sorted(seen) or "none"}')
    sys.exit()

for name in wb.sheetnames:
    if arg.lower() not in name.lower():
        continue
    ws = wb[name]
    print(f'########## {name}  ({ws.max_row} rows x {ws.max_column} cols) ##########')
    for row in ws.iter_rows():
        line = []
        for cell in row:
            if cell.value is None:
                continue
            marks = fmt_marks(cell)
            suffix = f'  [{", ".join(marks)}]' if marks else ''
            line.append(f'{cell.coordinate}: {cell.value}{suffix}')
        if line:
            print('\n'.join(line))
            print('-')
